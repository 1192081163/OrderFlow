from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook

import extract


def save_workbook(wb: Workbook, path: Path) -> Path:
    wb.save(path)
    return path


def test_material_code_rounds_0_55_galv_to_0_6g() -> None:
    assert extract.material_code("0.55mm Galv") == "0.6G"


def test_material_code_rounds_0_55_colorbond_colours_to_0_6cb() -> None:
    assert extract.material_code("0.55mm Monument") == "0.6CB"
    assert extract.material_code("0.55mm Surfmist") == "0.6CB"


def test_previous_business_day_uses_wa_weekday_not_china_may_holiday() -> None:
    assert extract.previous_business_day(dt.date(2026, 5, 5)) == dt.date(2026, 5, 4)


def test_previous_business_day_skips_wa_day_public_holiday() -> None:
    assert extract.previous_business_day(dt.date(2026, 6, 2)) == dt.date(2026, 5, 29)


def extracted_row(job: str, source_file: str, marker: str) -> extract.ExtractedRow:
    row = extract.ExtractedRow(source_file=source_file)
    row.values[6] = job
    row.values[19] = marker
    return row


def test_duplicate_jobs_keep_highest_source_version(tmp_path: Path) -> None:
    older_path = tmp_path / "29698__0178__old__29698 BEYOND RES SPLIT + CS.xlsx"
    newer_path = tmp_path / "29698__0216__new__29698 BEYOND RES SPLIT + CS.xlsx"
    older_path.touch()
    newer_path.touch()
    older = extracted_row("29698", older_path.name, "old")
    newer = extracted_row("29698", newer_path.name, "new")

    rows = extract.dedupe_latest_rows([older, newer], [older_path, newer_path])

    assert rows == [newer]


def test_duplicate_jobs_without_source_version_use_latest_file_mtime(tmp_path: Path) -> None:
    older_path = tmp_path / "29782 APEX MKD + ST.xlsx"
    newer_path = tmp_path / "29782 APEX SPLIT.xlsx"
    older_path.touch()
    newer_path.touch()
    older_time = 1_700_000_000
    newer_time = 1_700_000_100
    older_path.touch()
    newer_path.touch()
    import os

    os.utime(older_path, (older_time, older_time))
    os.utime(newer_path, (newer_time, newer_time))
    older = extracted_row("29782", older_path.name, "old")
    newer = extracted_row("29782", newer_path.name, "new")

    rows = extract.dedupe_latest_rows([older, newer], [older_path, newer_path])

    assert rows == [newer]


def worksheet_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    wb.create_sheet("Data")
    ws["C1"] = 99999
    ws["C2"] = "Test Builder"
    ws["C5"] = "2026-06-15"
    headers = [
        "Material",
        "Stock",
        "Qty",
        "Profile",
        "B/O",
        "Reveal Height",
        "Reveal Width",
        "Hand",
        "Hinge Qty",
        "Hinge Type",
        "Striker Type",
        "Striker Height",
        "Sill",
        "Slider",
        "Double",
        "CL1",
        "CL3",
        "Striker Type2",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(9, col).value = value
    return wb


def test_standard_worksheet_plain_hinge_stays_in_v_bucket(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Split",
        "85-125",
        2060,
        923,
        "RIGHT",
        "3",
        "HINGE PREP",
        "S1+RDL",
        "1000",
        "NO",
        "NO",
        "NO",
        "Split",
        "85-125",
        "S1+S1",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "plain.xlsx"), infer_manual=True)

    assert row.values[19:24] == [2, 5, 5, None, None]


def test_standard_worksheet_screw_fixed_prep_uses_w_bucket_and_skips_other_rows(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    rows = {
        11: [
            "1.05mm Zincanneal",
            "",
            1,
            "Modern",
            "95",
            2360,
            823,
            "LEFT WC",
            "3",
            "SCREW FIXED PREP",
            "S1+ZANDA 10421",
            "1000 + 1150",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            1,
        ],
        12: [
            "1.05mm Zincanneal",
            "",
            1,
            "Modern",
            "95",
            2360,
            823,
            "LEFT",
            "3",
            "SCREW FIXED PREP",
            "S1",
            "1000",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            1,
        ],
        14: [
            "Other",
            "",
            1,
            "Single Elec with Lock & View (Non-Rebated)",
            None,
            0,
            0,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            None,
            None,
            1,
        ],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "screw-fixed.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "MODERN"]
    assert row.values[19:24] == [2, 9, 3, 6, None]


def test_standard_worksheet_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Modern",
        "95",
        2060,
        823,
        "RIGHT",
        "2",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Modern",
        "95",
        "S1",
    ]
    for row_idx in (11, 12):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[12].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-worksheet.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "MODERN"]
    assert row.values[19:24] == [1, 3, 3, None, None]


def test_standard_worksheet_negative_qty_does_not_cancel_goods_qty(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    for row_idx, qty in ((11, -1), (12, 1)):
        values = [
            "1.05mm Zincanneal",
            "",
            qty,
            "Modern",
            "95",
            2060,
            823,
            "RIGHT",
            "2",
            "WELDED",
            "S1",
            "1000",
            "NO",
            "NO",
            "NO",
            "Modern",
            "95",
            "S1",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "negative-qty.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "MODERN"]
    assert row.values[19] == 1


def test_standard_worksheet_kd_parts_go_to_x_bucket_not_v_bucket(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        2,
        "Modern Knock Down",
        "114",
        2060,
        823,
        "RIGHT",
        "3",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Modern Knock Down",
        "114",
        "S1",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "worksheet-modern-kd.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "KD"]
    assert row.values[19:24] == [0, 11, 8, None, 8]


def test_standard_worksheet_equal_qty_keeps_source_goods_order(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    split_values = [
        "1.05mm Zincanneal",
        "",
        8,
        "Split",
        "85-125",
        2060,
        823,
        "RIGHT",
        "3",
        "WELDED",
        "S1",
        "1000",
        "NO",
        "NO",
        "NO",
        "Split",
        "85-125",
        "S1",
    ]
    for col, value in enumerate(split_values, start=1):
        ws.cell(11, col).value = value
    ws.cell(13, 1).value = "Cavity Sliders"
    cavity_values = [
        "",
        "",
        8,
        "Modern",
        "114",
        2060,
        700,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    for col, value in enumerate(cavity_values, start=1):
        ws.cell(14, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "equal-goods-order.xlsx"), infer_manual=True)

    assert row.values[10:14] == [8, "SPLIT", 8, "CS"]
    assert row.values[19] == 128


def test_standard_worksheet_cavity_soft_closer_does_not_count_as_cs_goods(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    ws.cell(12, 1).value = "Cavity Sliders"
    rows = {
        13: ["", "", 1, "Modern", "114", 2360, 700, "", "", "", "", "", "", "", "", "", "", ""],
        14: ["", "", 1, "Brio Soft Closer (bev to deliver)", "", 0, "", "", "", "", "", "", "", "", "", "", "", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "cavity-soft-closer.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "CS"]
    assert row.values[19] == 14


def sheet1_profile_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Job 30095"
    headers = ["QTY", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "HINGES", "TYPE", "HEIGHT", "HOLES"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    return wb


def test_sheet1_delivery_address_splits_zone_prefix_from_same_cell(tmp_path: Path) -> None:
    cases = [
        ("42b 28 SIGNAL TCE, COCKBURN TCE", "42B", "28 SIGNAL TCE, COCKBURN TCE"),
        ("03C, 62 CLAYTON STREET BELLEVUE", "03C", "62 CLAYTON STREET BELLEVUE"),
    ]

    for raw_address, expected_zone, expected_address in cases:
        wb = sheet1_profile_book()
        ws = wb["Sheet1"]
        ws["A5"] = "Delivery Address"
        ws["B5"] = raw_address
        ws.cell(13, 1).value = 1
        ws.cell(13, 2).value = "Modern"

        row = extract.extract_workbook(save_workbook(wb, tmp_path / f"{expected_zone}.xlsx"), infer_manual=True)

        assert row.values[4] == expected_zone
        assert row.values[5] == expected_address


def test_sheet1_hinges_quantity_goes_to_v_bucket(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, "12 + 6"]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "SPLIT"]
    assert row.values[19:24] == [2, 5, 4, None, None]


def test_sheet1_csk_dtna_counts_two_parts_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "CSK"
    ws.cell(11, 6).value = "DTNA"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-csk-dtna.xlsx"), infer_manual=True)

    assert row.values[21] == 20
    assert row.values[20] == 20


def test_sheet1_csk_dyna_tube_counts_four_parts_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "CSK"
    ws.cell(11, 6).value = "DYNA AND TUBE"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-csk-dyna-tube.xlsx"), infer_manual=True)

    assert row.values[21] == 40
    assert row.values[20] == 40


def test_sheet1_tradition_dyna_counts_one_part_per_quantity(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 6).value = "TRADITION"
    ws.cell(11, 6).value = "DYNA"
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", "10 (5 EACH SIDE)", "", "", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-tradition-dyna.xlsx"), infer_manual=True)

    assert row.values[21] == 10
    assert row.values[20] == 10


def test_sheet1_hinge_plates_quantity_goes_to_w_bucket(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(12, 7).value = "HINGE PLATES"
    values = [2, "Split A + B", 2060, 923, "RIGHT", 3, "Suit 100x75x2.5", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-hinge-plates.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "SPLIT"]
    assert row.values[19:24] == [4, 11, 2, 6, None]


def test_sheet1_profile_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "Split 180B/O (35 Door)", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, "12 + 6"]
    for row_idx in (13, 14):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[14].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-sheet1.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "SPLIT"]
    assert row.values[19:24] == [2, 5, 4, None, None]


def test_sheet1_profile_code_maps_to_commercial_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "144B/O", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-commercial-code.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_sheet1_cavity_slider_contributes_fourteen_mitres_each(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    ws.cell(10, 1).value = "Cavity Sliders"
    values = [2, "Modern", 2060, 700, "", "", "", "", "", ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-cavity-slider.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "CS"]
    assert row.values[19] == 28


def test_sheet1_numeric_letter_profile_code_maps_to_commercial_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "150B", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-commercial-number-letter.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_sheet1_kd_abbreviation_maps_to_kd_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [1, "144mm B/O KD - Screw Fix Mitre", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-kd-code.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]
    assert row.values[19:24] == [0, 6, 4, None, 4]


def test_sheet1_head_only_replacement_rows_do_not_count_as_kd_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    rows = {
        13: [1, "125mm B/O MODERN KNOCK DOWN-SCREW FIXED MITRE", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""],
        14: ["REPLACEMENT HEAD # 1", "112mm B/O MODERN KNOCK DOWN-SCREW FIXED MITRE HEAD ONLY", "", 723, "LEFT", "", "", "", "", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-head-only.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]


def test_standard_worksheet_head_only_order_still_counts_as_kd_goods(tmp_path: Path) -> None:
    wb = worksheet_book()
    ws = wb["Worksheet"]
    values = [
        "1.05mm Zincanneal",
        "",
        1,
        "Modern Knockdown - screw fix HEAD ONLY",
        "114",
        0,
        1315,
        "",
        "",
        "",
        "",
        "",
        "NO",
        "NO",
        "NO",
        "",
        "",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(11, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "head-only-order.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "KD"]


def test_sheet1_service_part_maps_to_part_goods(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [50, "Service Part - Hinge Plate Suit Fire Door", 2060, 923, "RIGHT", 3, "100X75X1.6", "S1", 1000, ""]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "sheet1-service-part.xlsx"), infer_manual=True)

    assert row.values[10:12] == [50, "PART"]


def main_sheet_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    wb.create_sheet("Profiles")
    ws["A1"] = "Job No 30354"
    ws["B2"] = "Fire Door Maintenance"
    headers = ["Door #", "PROFILE", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "GUARDS", "TYPE", "HEIGHT", "BOLT", "BRICK TIES"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    return wb


def test_main_sheet_commercial_widths_over_threshold_count_oversize(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    rows = {
        13: ["D.4", "A", 2410, 1920, "DOUBLE", "8 (4 EACH SIDE)", "100X100X2.5", "YES", "-", "-", "-", 10],
        14: ["D.3", "A", 2110, 1030, "RIGHT", 4, "100X100X2.5", "YES", "S1", 1020, "-", 8],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profile.xlsx"), infer_manual=True)

    assert row.values[7] == 2
    assert row.values[10:12] == [2, "COMMERCIAL"]
    assert row.values[19:24] == [3, 18, 1, 12, None]


def test_main_sheet_offset_hand_column_still_counts_oversize_width(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = [
        "Door #",
        "(PER A_6020 & A_6021)",
        "PROFILE",
        "THICKNESS",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "QTY",
        "TO SUIT",
        "TYPE",
        "HEIGHT",
        "BOLT",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "ON.LG035B",
        "13",
        "CUSTOM",
        "40mm",
        2400,
        1856,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-offset-hand.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_main_sheet_profileless_table_extracts_commercial_hardware(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "MAIL ROOM",
        "SPLIT 85-125B/O ",
        "40mm",
        2060,
        923,
        "RIGHT",
        4,
        "100X75X2.5",
        "MORTICE LOCK (S1)",
        1032,
        "12 + 6",
        "NOT REQUIRED",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 5, 1, 4, None]


def test_main_sheet_profileless_width_over_threshold_counts_oversize(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "D2.01/S038.1",
        "DMF02",
        "DL01.M = 40mm Thick",
        2380,
        1246,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless-oversize.xlsx"), infer_manual=True)

    assert row.values[7] == 1


def test_main_sheet_profileless_hand_header_adds_commercial_double_mitre(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "D2.01/S038.1",
        "DMF02",
        "DL01.M = 40mm Thick",
        2380,
        1246,
        "DOUBLE",
        "8 (4 EACH SIDE)",
        "100X75X2.5",
        "-",
        "-",
        "8 (4 EACH JAMB)",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profileless-double.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19] == 2


def test_main_sheet_double_action_commercial_adds_half_mitre_each(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = [
        "Door #",
        "(PER A_6020 & A_6021)",
        "PROFILE",
        "THICKNESS",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "DOUBLE ACTION BOXES",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    for row_idx in (13, 14):
        values = [f"ON.L{row_idx}", "08A", "CUSTOM D/A", "40mm", 2260, 2272, "DOUBLE ACTION", "2 - WELDED IN TO HEAD OF FRAME"]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-double-action.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "COMMERCIAL"]
    assert row.values[19] == 3


def test_main_sheet_profileless_hidden_detail_rows_are_ignored(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    headers = ["Door #", "TYPE", "THICKNESS", "HEIGHT", "WIDTH", "HAND", "QTY", "TO SUIT", "TYPE", "HEIGHT", "HOLES", "BRACKETS"]
    for col, value in enumerate(headers, start=1):
        ws.cell(12, col).value = value
    values = [
        "MAIL ROOM",
        "SPLIT 85-125B/O ",
        "40mm",
        2060,
        923,
        "RIGHT",
        4,
        "100X75X2.5",
        "MORTICE LOCK (S1)",
        1032,
        "12 + 6",
        "NOT REQUIRED",
    ]
    for row_idx in (13, 14):
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
    ws.row_dimensions[14].hidden = True

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "hidden-main-profileless.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]
    assert row.values[19:24] == [1, 5, 1, 4, None]


def test_main_sheet_profile_hash_header_uses_one_product_when_qty_is_hardware(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    for col in range(1, 13):
        ws.cell(12, col).value = None
    ws.cell(15, 7).value = "HINGE"
    headers = [
        "Door #",
        "PROFILE #",
        "TYPE",
        "HEIGHT",
        "WIDTH",
        "HAND",
        "QTY",
        "SIZE",
        "TYPE",
        "HEIGHT",
        "Switch",
        "Switch Hole",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(16, col).value = value
    values = ["00-059-D1", "CUSTOM D1", "DLFRCL", 2410, 1062, "RIGHT", 5, "100X100X3.2", "S1", 1032, "1 IN HEAD", "NO"]
    for col, value in enumerate(values, start=1):
        ws.cell(17, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "main-profile-hash.xlsx"), infer_manual=True)

    assert row.values[10:12] == [1, "COMMERCIAL"]


def test_main_sheet_split_part_only_counts_one_mitre_each(tmp_path: Path) -> None:
    wb = sheet1_profile_book()
    ws = wb["Sheet1"]
    values = [
        8,
        "CUSTOM SPLIT PART B ONLY (30MM ARCHITRAVE)",
        2060,
        910,
        "RIGHT",
        8,
        "100X75X2.5",
        "S1",
        1000,
        "-",
        "",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(13, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "split-part-b-only.xlsx"), infer_manual=True)

    assert row.values[10:12] == [8, "SPLIT"]
    assert row.values[19] == 8


def test_nonstandard_worksheet_door_qty_rows_extract_goods_qty(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws["A1"] = "Ausmet Job # 29463"
    ws["C2"] = "Carnarvon Timber & Hardware"
    ws["C5"] = "2026-05-13"
    headers = [
        "Material",
        "Door #",
        "B/O",
        "Qty",
        "WALL TYPE",
        "FRAME TYPE",
        "DOOR THICKNESS",
        "REVEAL HEIGHT",
        "REVEAL WIDTH",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(10, col).value = value
    rows = {
        11: ["1.05mm Galv", "", "", 0, "", "", "", "", ""],
        12: ["1.05mm Galv", "AC BOX SPLIT A & B", "", 1, "", "", "", 560, 770],
        13: ["1.05mm Galv", "AC BOX SPLIT A & B", "", 1, "", "", "", 575, 700],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "nonstandard-split.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "SPLIT"]
    assert row.values[19] == 4


def test_door_skins_capping_rows_do_not_fallback_to_commercial_goods(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    ws["A1"] = "AUMSET JOB # 29322"
    ws["B2"] = "Australian Fire Door Company"
    headers = ["MATERIAL", "PROFILE", "QUANTITY", "WIDTH", "LENGTH", "BOLT", ""]
    for col, value in enumerate(headers, start=1):
        ws.cell(11, col).value = value
    rows = {
        12: ["0.55mm Deep Ocean", "A", 2, 845, 2030, "", "FLAT SHEET"],
        13: ["0.55mm Deep Ocean", "B", 2, "", 2100, "", "CAPPING"],
        14: ["0.55mm Deep Ocean", "B", 2, "", 1000, "", "CAPPING"],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "AFDC Door Skins.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "DS"]


def test_trad_dyna_split_profile_counts_as_commercial_goods(tmp_path: Path) -> None:
    wb = main_sheet_book()
    ws = wb["Main Sheet"]
    rows = {
        13: ["D8", "A", 2075, 1011, "LEFT", 4, "100x100x2.5", "S1", 1030, "", ""],
        20: ["D7", "125BO SPLIT", 2060, 910, "LEFT", 4, "100X100X2.5", "S1", 1000, "-", ""],
    }
    for row_idx, values in rows.items():
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value

    row = extract.extract_workbook(save_workbook(wb, tmp_path / "AFDC Trad Dyna Single Rebate.xlsx"), infer_manual=True)

    assert row.values[10:12] == [2, "COMMERCIAL"]
